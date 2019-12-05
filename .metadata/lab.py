from openpyxl import load_workbook
from openpyxl import workbook

def is_del(data):#判断是否有%
    if data==None:
        print('del1')
        return True
    l=len(data)
    for i in range(l):
        if data[i]=='%':
            print('keep')
            return False
        if data[i]=='=':
            print('del2')
            return True
    if i==l-1:
        print('del3')
        return False

wb=load_workbook('南航深航航段客座率-模板.xlsx')
ws=wb['粘贴板']
i=2
rowsum=ws.max_row
while True:
    if is_del(ws.cell(row=i, column=2).value):
        ws.delete_cols(i)
    if i>=rowsum:
        break
    i=i+1
wb.save('new.xlsx')


