from openpyxl import load_workbook
from openpyxl import Workbook
from onedaydata import processdata
wobook1={0:'昨天CZ',1:'今天CZ',2:'明天CZ',3:'D2CZ',4:'D3CZ',5:'D4CZ'}
wobook2={6:'昨天ZH',7:'今天ZH',8:'明天ZH',9:'D2ZH',10:'D3ZH',11:'D4ZH'}
result1=Workbook
result2=Workbook
#第几张sheet
j=2#某张sheet第几行
bokshet=0#读取sheet和workbook判断
for data in processdata():
    if bokshet==0 or bokshet==1 or bokshet==2 or bokshet==6 or bokshet==7 or bokshet==8:
        i=1
        if bokshet==6 or bokshet==7 or bokshet==8:
            wb1=load_workbook('南航深航航段客座率-模板.xlsx')
            ws = wb1[wobook2[bokshet]]
        else:
            wb1 = load_workbook('南航深航航段客座率-模板.xlsx')
            ws = wb1[wobook1[bokshet]]
    else:
        i=2
        if bokshet==3 or bokshet==4 or bokshet==5:
            wb2=load_workbook('南航深航航段客座率-模板(2-4天).xlsx')
            ws=wb2[wobook1[bokshet]]
        else:
            wb2 = load_workbook('南航深航航段客座率-模板(2-4天).xlsx')
            ws = wb2[wobook2[bokshet]]
    for k in range(len(data)):
        ws.cell(row=j, column=1).value = data[k]
        j = j + 1
    j=2
    print(data)
    if bokshet>=11:
        break
    else:
        bokshet=bokshet+1
    if i==1:
        wb1.save('南航深航航段客座率-模板.xlsx')
    else:
        wb2.save('南航深航航段客座率-模板(2-4天).xlsx')


