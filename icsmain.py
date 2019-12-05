import pyautogui
import time
import datetime
import scratch
from openpyxl import load_workbook
from openpyxl import Workbook
from onedaydata import processdata
import openpyxl
wobook1={0:'昨天CZ',1:'今天CZ',2:'明天CZ',3:'D2CZ',4:'D3CZ',5:'D4CZ'}
wobook2={6:'昨天ZH',7:'今天ZH',8:'明天ZH',9:'D2ZH',10:'D3ZH',11:'D4ZH'}
result1=Workbook
result2=Workbook
COMPANYS=['cz','zh']
pyautogui.doubleClick(x=880,y=490,button='left')
pyautogui.typewrite('230186\t650824a@\n',0.1)
time.sleep(8)
pyautogui.typewrite('\n')
pyautogui.typewrite('si:f/28302/65824w',0.1)
pyautogui.click(x=218,y=60)
time.sleep(5)
#增加判断一个是否连接
for k in range(2):
    if k==0:
        company='cz'
    else:
        company='zh'
    if pyautogui.pixelMatchesColor(218,60,(50,160,79)):
        pyautogui.click(x=218,y=60)
        pyautogui.moveTo(15,220,duration=0.5)
        i=datetime.datetime.now()
        d=i.day
        m=i.month
        y=i.year
        for j in range (6):
            date=scratch.timeprocess(d-1,m,y)
            print(date)
            if company=='cz':
                pyautogui.typewrite('fdl /k/c/'+date+'/szx/'+company+'/99',0.15)
            else:
                pyautogui.typewrite('fdl /k/d/' + date + '/szx/' + company + '/99', 0.15)
            pyautogui.click(x=218,y=60)
            while True:
                #加一个剪切板操作判断
                if pyautogui.locateOnScreen('newend.png')or pyautogui.locateOnScreen('end.png') or pyautogui.locateOnScreen('end_.png'):
                    pyautogui.hotkey('ctrlleft','pageup')
                    break
                else:
                    pyautogui.typewrite("pn",0.3)
                    pyautogui.click(x=218,y=60)
            d=d+1

j=2#某张sheet第几行
bokshet=0#读取sheet和workbook判断
for data in processdata():
    if bokshet==0 or bokshet==1 or bokshet==2 or bokshet==6 or bokshet==7 or bokshet==8:
        i=1
        if bokshet==6 or bokshet==7 or bokshet==8:
            wb1=load_workbook('南航深航航段客座率-模板.xlsx')
            ws = wb1[wobook2[bokshet]]
        else:
            wb1 = load_workbook('南航深航航段客座率-模板.xlsx')
            ws = wb1[wobook1[bokshet]]
    else:
        i=2
        if bokshet==3 or bokshet==4 or bokshet==5:
            wb2=load_workbook('南航深航航段客座率-模板(2-4天).xlsx')
            ws=wb2[wobook1[bokshet]]
        else:
            wb2 = load_workbook('南航深航航段客座率-模板(2-4天).xlsx')
            ws = wb2[wobook2[bokshet]]
    for k in range(len(data)):
        ws.cell(row=j, column=1).value = data[k]
        j = j + 1
    j=2
    print(data)
    if bokshet>=11:
        break
    else:
        bokshet=bokshet+1
    if i==1:
        wb1.save('南航深航航段客座率-模板.xlsx')
    else:
        wb2.save('南航深航航段客座率-模板(2-4天).xlsx')

#新建excel
wb2=openpyxl.Workbook()
wb2.save('结果1.xlsx')
print('新建成功')

#读取数据
wb1=openpyxl.load_workbook('南航深航航段客座率-模板.xlsx')
wb2=openpyxl.load_workbook('结果1.xlsx')
sheet1=wb1['结果']
sheet2=wb2.create_sheet('结果')

max_row=sheet1.max_row#最大行数
max_column=sheet1.max_column#最大列数

for m in range(1,max_row+1):
     for n in range(97,97+max_column):#chr(97)='a'
        n=chr(n)#ASCII字符
        i='%s%d'%(n,m)#单元格编号
        cell1=sheet1[i].value#获取data单元格数据
        sheet2[i].value=cell1#赋值到test单元格

wb2.save('结果1.xlsx')#保存数据
wb1.close()#关闭excel
wb2.close()

wb2=openpyxl.Workbook()
wb2.save('结果1.xlsx')
print('新建成功')

#读取数据
wb1=openpyxl.load_workbook('南航深航航段客座率-模板(2-4天).xlsx')
wb2=openpyxl.load_workbook('结果2.xlsx')
sheet1=wb1['结果']
sheet2=wb2.create_sheet('结果')

max_row=sheet1.max_row#最大行数
max_column=sheet1.max_column#最大列数

for m in range(1,max_row+1):
     for n in range(97,97+max_column):#chr(97)='a'
        n=chr(n)#ASCII字符
        i='%s%d'%(n,m)#单元格编号
        cell1=sheet1[i].value#获取data单元格数据
        sheet2[i].value=cell1#赋值到test单元格

wb2.save('结果2.xlsx')#保存数据
wb1.close()#关闭excel
wb2.close()

