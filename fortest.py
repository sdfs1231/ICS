from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl

#新建excel
wb2=openpyxl.Workbook()
wb2.save('test.xlsx')
print('新建成功')

#读取数据
wb1=openpyxl.load_workbook('南航深航航段客座率-模板.xlsx')
wb2=openpyxl.load_workbook('test.xlsx')
sheet1=wb1['结果']
sheet2=wb2.create_sheet('结果')

max_row=sheet1.max_row#最大行数
max_column=sheet1.max_column#最大列数

for m in range(1,max_row+1):
     for n in range(97,97+max_column):#chr(97)='a'
        n=chr(n)#ASCII字符
        i='%s%d'%(n,m)#单元格编号
        cell1=sheet1[i].value#获取data单元格数据
        sheet2[i].value=cell1#赋值到test单元格

wb2.save('test.xlsx')#保存数据
wb1.close()#关闭excel
wb2.close()