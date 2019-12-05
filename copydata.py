from openpyxl import load_workbook
from openpyxl import Workbook
from onedaydata import processdata
wb=Workbook()
i=1
j=2
for data in processdata():
        ws = wb.create_sheet('cz' + 'i')
        for k in range(len(data)):
            ws.cell(row=j, column=1).value = data[k]
            j = j + 1
        j=2
        print(data)
        i=i+1

wb.save('new1.xlsx')